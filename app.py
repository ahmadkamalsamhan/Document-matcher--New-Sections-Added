import streamlit as st
import pandas as pd
import re
import tempfile
import os
import time
from openpyxl import load_workbook

st.set_page_config(page_title="📊 Nesma & Partners - Document Processing App", layout="wide")
st.title("📊 Nesma & Partners - Document Processing App ")

# -----------------------------
# GLOBAL RESET BUTTON
# -----------------------------
if st.button("🗑 Clear/Reset Entire App"):
    keys_to_clear = ["uploaded_files", "tmp_path", "filter_file"]
    cleared = False
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
            cleared = True
    if cleared:
        st.success("✅ App fully reset. All uploaded files and filters cleared.")
        st.experimental_rerun()
    else:
        st.success("✅ App is already clean. You can continue normally.")
        
# -----------------------------
# PART 1 - MATCHING
# -----------------------------
st.header("Matching Two Excel Files")

uploaded_files = st.file_uploader(
    "Upload Excel files", type="xlsx", accept_multiple_files=True, key="uploaded_files"
)

if uploaded_files:
    st.subheader("Select files to use for matching")
    selected_files = [f for f in uploaded_files if st.checkbox(f.name, value=True)]
    if len(selected_files) >= 2:
        st.success(f"{len(selected_files)} files selected for matching.")

        df1_columns = pd.read_excel(selected_files[0], nrows=0).columns.tolist()
        df2_columns = pd.read_excel(selected_files[1], nrows=0).columns.tolist()

        st.subheader("Step 1: Select column to match")
        match_col1 = st.selectbox(f"Column from {selected_files[0].name}", df1_columns)
        match_col2 = st.selectbox(f"Column from {selected_files[1].name}", df2_columns)

        st.subheader("Step 2: Select additional columns to include in the result")
        include_cols1 = st.multiselect(f"Columns from {selected_files[0].name}", df1_columns)
        include_cols2 = st.multiselect(f"Columns from {selected_files[1].name}", df2_columns)

        # --- NEW: Matching Mode Selection ---
        match_mode = st.radio(
            "Select Matching Mode",
            [
                "Mode 1 – All Logics",
                "Mode 2 – Structured Code Extraction (ICT-DP-PS2-22D-5)",
                "Mode 3 – BV/FH/CM/ND Normalization (BV-01, FH-03, CM-50, ND 53-CP, etc.)"
            ]
        )

        if st.button("Step 3: Start Matching"):
            if not match_col1 or not match_col2:
                st.warning("⚠️ Please select columns to match.")
            elif not include_cols1 and not include_cols2:
                st.warning("⚠️ Please select at least one additional column to include in the result.")
            else:
                st.info(f"⏳ Matching in progress using {match_mode} ...")

                try:
                    df1_small = pd.read_excel(selected_files[0], usecols=[match_col1] + include_cols1)
                    df2_small = pd.read_excel(selected_files[1], usecols=[match_col2] + include_cols2)

                    # ==============================
                    # MODE 1: ORIGINAL LOGIC (untouched)
                    # ==============================
                    if match_mode.startswith("Mode 1"):

                        def normalize(text):
                            if pd.isna(text): return ""
                            text = str(text).lower()
                            text = re.sub(r'[^a-z0-9\s]', ' ', text)
                            text = re.sub(r'\s+', ' ', text).strip()
                            return text

                        df1_small['token_set'] = df1_small[match_col1].apply(normalize).str.split().apply(set)
                        df2_small['norm_match'] = df2_small[match_col2].apply(normalize)

                    # ==============================
                    # MODE 2: STRUCTURED CODE EXTRACTION
                    # ==============================
                    elif match_mode.startswith("Mode 2"):

                        def extract_code(text):
                            if pd.isna(text):
                                return ""
                            text = str(text)
                            m = re.search(r'[A-Za-z]{2,}-[A-Za-z0-9-]+', text)
                            return m.group(0).lower() if m else text.lower()

                        df1_small['norm_match'] = df1_small[match_col1].apply(extract_code)
                        df2_small['norm_match'] = df2_small[match_col2].apply(extract_code)

                    # ==============================
                    # MODE 3: BV/FH/CM/ND Normalization
                    # ==============================
                    elif match_mode.startswith("Mode 3"):

                        def normalize_bv_fh(text):
                            if pd.isna(text):
                                return ""
                            text = str(text).strip().lower()
                            text = re.sub(r'[_ ]', '-', text)  # unify separators
                            text = re.sub(r'branch to', '', text, flags=re.IGNORECASE)
                            # Match BV, FH, CM, ND codes with optional numbers and suffixes
                            pattern = r'\b(bv-?\d+|fh-?\d+|cm-?\d+|nd-?\d+(?:-cp)?)\b'
                            matches = re.findall(pattern, text)
                            if matches:
                                return " / ".join(matches)  # join multiple codes if present
                            return text

                        df1_small['norm_match'] = df1_small[match_col1].apply(normalize_bv_fh)
                        df2_small['norm_match'] = df2_small[match_col2].apply(normalize_bv_fh)

                    # ==============================
                    # MATCHING ENGINE (shared across all modes)
                    # ==============================
                    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    tmp_path = tmp_file.name
                    tmp_file.close()
                    pd.DataFrame(columns=include_cols1 + include_cols2).to_excel(tmp_path, index=False)

                    total_rows = len(df2_small)
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    start_time = time.time()

                    batch_size = 200
                    buffer_rows = []

                    for idx, row in df2_small.iterrows():
                        norm_val = row['norm_match']
                        if not norm_val:
                            continue

                        if match_mode.startswith("Mode 1"):
                            row_tokens = set(norm_val.split())
                            mask = df1_small['token_set'].apply(lambda x: row_tokens.issubset(x))
                        else:
                            mask = df1_small['norm_match'] == norm_val

                        matched_rows = df1_small.loc[mask, include_cols1].copy()
                        if not matched_rows.empty:
                            for col in include_cols2:
                                matched_rows[col] = row[col]
                            buffer_rows.append(matched_rows)

                        if len(buffer_rows) >= batch_size:
                            batch_df = pd.concat(buffer_rows, ignore_index=True)
                            with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='a',
                                                if_sheet_exists='overlay') as writer:
                                startrow = writer.sheets['Sheet1'].max_row
                                batch_df.to_excel(writer, index=False, header=False, startrow=startrow)
                            buffer_rows = []

                        progress_bar.progress((idx + 1) / total_rows)
                        status_text.text(f"Processing row {idx + 1}/{total_rows} ({(idx + 1) / total_rows * 100:.1f}%)")

                    # Final flush of remaining buffer
                    if buffer_rows:
                        batch_df = pd.concat(buffer_rows, ignore_index=True)
                        with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='a',
                                            if_sheet_exists='overlay') as writer:
                            startrow = writer.sheets['Sheet1'].max_row
                            batch_df.to_excel(writer, index=False, header=False, startrow=startrow)

                    # ==============================
                    # PREPARE MATCHED & UNMATCHED
                    # ==============================
                    matched_df = pd.read_excel(tmp_path)

                    # Detect unmatched rows from df1_small
                    if match_mode.startswith("Mode 1"):
                        matched_tokens_list = matched_df[match_col1].apply(lambda x: set(str(x).lower().split()))
                        df1_tokens = df1_small['token_set']
                        mask_unmatched = ~df1_tokens.apply(lambda x: any(x == mt for mt in matched_tokens_list))
                    else:
                        matched_values = matched_df['norm_match'].unique()
                        mask_unmatched = ~df1_small['norm_match'].isin(matched_values)

                    unmatched_df = df1_small.loc[mask_unmatched, include_cols1]

                    end_time = time.time()
                    st.success(f"✅ Matching complete in {end_time - start_time:.2f} seconds")

                    # ==============================
                    # SHOW RESULTS & DOWNLOAD
                    # ==============================
                    st.subheader("Preview of Matched Results (first 100 rows)")
                    st.dataframe(matched_df.head(100))

                    st.subheader("Preview of Unmatched Rows (first 100 rows)")
                    st.dataframe(unmatched_df.head(100))

                    # Download matched
                    with open(tmp_path, "rb") as f:
                        st.download_button("💾 Download Matched Results", data=f,
                                           file_name="matched_results.xlsx")

                    # Download unmatched
                    tmp_unmatched = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    unmatched_df.to_excel(tmp_unmatched.name, index=False)
                    with open(tmp_unmatched.name, "rb") as f:
                        st.download_button("💾 Download Unmatched Results", data=f,
                                           file_name="unmatched_results.xlsx")
                    os.remove(tmp_unmatched.name)
                    os.remove(tmp_path)

                except Exception as e:
                    st.error(f"❌ Error during matching: {e}")

    else:
        st.warning("⚠️ Please select at least 2 files for matching.")

# -----------------------------
# PART 2 - SEARCH & FILTER
# -----------------------------
st.header("Search & Filter Data")

uploaded_filter_file = st.file_uploader(
    "Upload an Excel file for filtering", type="xlsx", key="filter_file"
)

# helper function
def filter_dataframe_columnwise_partial(df, column_keywords, logic="AND"):
    masks = []
    for col, keywords in column_keywords.items():
        col_series = df[col].astype(str).str.lower()
        keyword_masks = [col_series.str.contains(k.lower().strip(), regex=False, na=False) for k in keywords]
        if keyword_masks:
            masks.append(pd.concat(keyword_masks, axis=1).any(axis=1))
        else:
            masks.append(pd.Series([True]*len(df), index=df.index))
    if logic.upper() == "AND":
        final_mask = pd.concat(masks, axis=1).all(axis=1)
    else:
        final_mask = pd.concat(masks, axis=1).any(axis=1)
    return final_mask

if uploaded_filter_file:
    df_filter = pd.read_excel(uploaded_filter_file)
    st.success(f"✅ File {uploaded_filter_file.name} uploaded with {len(df_filter)} rows.")

    search_all = st.checkbox("🔎 Search across all columns (ignore column selection)")

    column_keywords = {}
    col_logic = "AND"
    keywords_input = ""
    global_logic = "OR"

    # -----------------------------
    # NEW FEATURE: MULTI-KEYWORD SAME COLUMN
    # -----------------------------
    st.subheader("Search in Single Column (up to 5 keywords/sentences)")

    # Step 1: user selects the column
    same_col = st.selectbox("Select column for multi-keyword search", ["-- None --"] + df_filter.columns.tolist())

    same_col_keywords = []

    # Step 2: only show keyword inputs if a real column is selected
    if same_col != "-- None --":
        # Reset if column changes
        if "last_same_col" not in st.session_state or st.session_state.last_same_col != same_col:
            st.session_state.keyword_count = 1
            for k in list(st.session_state.keys()):
                if k.startswith("samecol_kw_"):
                    del st.session_state[k]
            st.session_state.last_same_col = same_col

        if "keyword_count" not in st.session_state:
            st.session_state.keyword_count = 1

        # Add field button
        if st.button("➕ Add another keyword (max 5)") and st.session_state.keyword_count < 5:
            st.session_state.keyword_count += 1

        # Dynamic keyword fields
        for i in range(st.session_state.keyword_count):
            val = st.text_input(f"Keyword {i+1} for '{same_col}'", key=f"samecol_kw_{i}")
            if val.strip():
                same_col_keywords.append(val.strip())

    # -----------------------------
    # EXISTING COLUMN-WISE SEARCH
    # -----------------------------
    if not search_all:
        filter_cols = st.multiselect("Select columns to apply filters on", df_filter.columns.tolist())
        for col in filter_cols:
            keywords = st.text_input(f"Keywords for '{col}' (comma-separated)")
            if keywords:
                column_keywords[col] = [k.strip() for k in keywords.split(",") if k.strip()]
        col_logic_radio = st.radio(
            "Select cross-column logic for multiple columns",
            options=["AND (match all columns)", "OR (match any column)"],
            index=0
        )
        col_logic = "AND" if col_logic_radio.startswith("AND") else "OR"

    # -----------------------------
    # EXISTING GLOBAL SEARCH
    # -----------------------------
    st.subheader("Global Search Options")
    keywords_input = st.text_input("Enter keywords to search across all columns (comma-separated)")
    global_logic_radio = st.radio(
        "Global keyword logic",
        options=["AND (match all keywords)", "OR (match any keyword)"],
        index=1
    )
    global_logic = "AND" if global_logic_radio.startswith("AND") else "OR"

    max_preview = st.number_input("Preview rows (max)", min_value=10, max_value=1000, value=200)

    if st.button("🔍 Apply Filter"):
        df_result = df_filter.copy()
        final_mask = pd.Series([True]*len(df_result), index=df_result.index)

        # NEW SAME-COLUMN SEARCH
        if same_col != "-- None --" and same_col_keywords:
            col_series = df_result[same_col].astype(str).str.lower()
            keyword_masks = [col_series.str.contains(k.lower(), regex=False, na=False) for k in same_col_keywords]
            same_col_mask = pd.concat(keyword_masks, axis=1).any(axis=1)
            final_mask &= same_col_mask

            # Add "Searched keyword" column (can hold multiple if needed)
            def match_keywords(val):
                matches = [kw for kw in same_col_keywords if kw.lower() in str(val).lower()]
                return ", ".join(matches) if matches else ""
            df_result["Searched keyword"] = df_result[same_col].apply(match_keywords)

        # EXISTING COLUMN-WISE FILTERING
        if not search_all and column_keywords:
            col_mask = filter_dataframe_columnwise_partial(df_result, column_keywords, col_logic)
            final_mask &= col_mask

        # EXISTING GLOBAL SEARCH
        if keywords_input.strip():
            keywords = [k.lower().strip() for k in keywords_input.split(",") if k.strip()]
            masks = []
            for k in keywords:
                mask = df_result.astype(str).apply(
                    lambda row: row.str.lower().str.contains(re.escape(k), na=False).any(),
                    axis=1
                )
                masks.append(mask)
            if global_logic.upper() == "AND":
                global_mask = pd.concat(masks, axis=1).all(axis=1)
            else:
                global_mask = pd.concat(masks, axis=1).any(axis=1)
            final_mask &= global_mask

        df_result = df_result[final_mask]

        if df_result.empty:
            st.error("❌ No rows matched your filters.")
        else:
            st.success(f"✅ Found {len(df_result)} matching rows.")
            st.dataframe(df_result.head(max_preview))

            csv = df_result.to_csv(index=False).encode("utf-8")
            st.download_button("💾 Download Filtered Results (CSV)", data=csv,
                               file_name="filtered_results.csv")

            tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            df_result.to_excel(tmp_xlsx.name, index=False)
            with open(tmp_xlsx.name, "rb") as f:
                st.download_button("💾 Download Filtered Results (XLSX)", data=f,
                                   file_name="filtered_results.xlsx")
            os.remove(tmp_xlsx.name)
# -----------------------------
# PART 3 - Gather Documents from Part 1 Result
# -----------------------------
st.header("Gather Documents - Groups")

# Step 0: Upload Part 1 result
part1_file = st.file_uploader("Upload Part 1 Result Excel file", type="xlsx", key="part3_file")

if part1_file:
    part3_df = pd.read_excel(part1_file)
    st.success(f"✅ Loaded uploaded file with {len(part3_df)} rows")

    # Step 1: Select key column to group/gather by
    key_col = st.selectbox("Select column to gather on (e.g., Start Structure)", part3_df.columns)

    # Step 2: Automatically select all other columns for result
    default_cols = [col for col in part3_df.columns if col != key_col]
    include_cols = st.multiselect(
        "Select additional columns to include in the result",
        part3_df.columns,
        default=default_cols
    )

    # Step 3: Start gathering
    if st.button("Start Part 3 Gathering"):
        if not key_col:
            st.warning("⚠️ Please select a key column.")
        elif not include_cols:
            st.warning("⚠️ Please select at least one additional column to include in the result.")
        else:
            st.info("⏳ Gathering in progress (memory-safe)...")
            import time
            start_time = time.time()
            progress_bar = st.progress(0)
            status_text = st.empty()

            results = []
            grouped = part3_df.groupby(key_col, dropna=False)
            total_groups = len(grouped)

            for idx, (group_val, group_df) in enumerate(grouped, 1):
                merged_row = {key_col: group_val}
                for col in include_cols:
                    merged_row[col] = " / ".join(group_df[col].astype(str).unique())
                results.append(merged_row)

                progress_bar.progress(idx / total_groups)
                status_text.text(f"Processing {idx}/{total_groups} groups ({idx/total_groups*100:.1f}%)")

            final_grouped = pd.DataFrame(results)
            end_time = time.time()
            st.success(f"✅ Gathering complete in {end_time - start_time:.2f} seconds")

            # Preview first 100 rows
            st.subheader("Preview of Gathered Results (first 100 rows)")
            st.dataframe(final_grouped.head(100))

            # Download full results
            import tempfile, os
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            final_grouped.to_excel(tmp_file.name, index=False)
            with open(tmp_file.name, "rb") as f:
                st.download_button("💾 Download Gathered Results", data=f,
                                   file_name="part3_gathered_results.xlsx")
            os.remove(tmp_file.name)
else:
    st.info("If needed Please upload the Excel file you got from Part 1 to start Part 3.")
